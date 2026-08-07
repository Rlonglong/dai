import functools
import re

import openpyxl
import pandas as pd

PATH_DATA_RULE = "/run/media/root/D/data/_0Data_Administrative/檔規/(cur)大數據平臺案_檔案儲存與來源(v4.11)對照_20260115.xlsx"

# ==========================================
# 特例設定：哪些表的哪些欄位，長度要改用「H_資料類型長度」(型別定義長度)
# 而不是「S_資料長度」(實際觀測長度)。
# 原本這些散落在一長串 if/elif table_name 判斷式裡，現在集中在這裡管理，
# 新增表格或欄位只要加一行 set，不用碰函式邏輯本身。
# ==========================================
USE_SPEC_LENGTH_COLUMNS: dict[str, set[str]] = {
    # 20250502新增NAME #20260115新增REL_1_TITLE、REL_2_TITLE、REL_3_TITLE-嘉偉
    "T_CUST": {"ADR", "RESIDENT_CITY_DIST", "NAME", "REL_1_TITLE", "REL_2_TITLE", "REL_3_TITLE"},
    "T_ATM_IN": {"POST_BRH_NAME", "ATM_ADR_CITY"},
    "T_TXN_PS": {"TXN_NAME"},
    "T_TXN_PS_INTEREST": {"TXN_NAME"},
    "T_TXN_PS_INTEREST_NORMAL": {"TXN_NAME"},
}

# 固定長度覆寫：不管檔規寫什麼，這個表的這個欄位長度就直接用這個值。
# 對應原本 T_ATM_OR / TXN_YM 那個寫死 8 的特例。
FIXED_LENGTH_OVERRIDE: dict[tuple[str, str], int] = {
    ("T_ATM_OR", "TXN_YM"): 8,
}


@functools.lru_cache(maxsize=1)
def _load_workbook_sheet_names() -> list[str]:
    """
    整份檔規的 sheet 名稱只在 process 生命週期內讀一次、快取起來。
    原本 openpyxl.load_workbook() 每次呼叫 data_rule() 都會重跑一次，
    如果同一個 process 裡要處理多張表（很常見），這筆成本會被重複付很多次。
    """
    wb = openpyxl.load_workbook(PATH_DATA_RULE, read_only=True)
    return wb.sheetnames


@functools.lru_cache(maxsize=1)
def _load_master_sheet() -> pd.DataFrame:
    """總表（table_name -> 分頁名稱 的對照表）也只讀一次、快取起來。"""
    sheet_names = _load_workbook_sheet_names()
    master_sheet = next((s for s in sheet_names if "總表" in s), None)
    if master_sheet is None:
        raise ValueError(f"❌ 檔規 {PATH_DATA_RULE} 裡找不到包含「總表」字樣的分頁")
    return pd.read_excel(PATH_DATA_RULE, sheet_name=master_sheet, usecols=[4, 5])


def _extract_spec_length(type_len_str) -> int:
    """從「H_資料類型長度」欄位（例如 "VARCHAR2(20)"）取出數字部分當長度"""
    match = re.search(r"(\d+)", str(type_len_str))
    if not match:
        raise ValueError(f"❌ 無法從 H_資料類型長度 解析出長度: {type_len_str!r}")
    return int(match.group(1))


def data_rule(target_table_name: str) -> list[list]:
    """
    從檔規 Excel 讀出 target_table_name 的欄位定義。

    回傳格式維持跟舊版一樣，是 [lengths, start_offsets, names] 三個 list：
      - lengths[i]：第 i 個欄位的長度
      - start_offsets[i]：第 i 個欄位在整筆定長資料裡的起始位移（0-based，
        故意比 lengths/names 多一個元素，最後一個值等於資料總長度）
      - names[i]：第 i 個欄位的名稱

    要組 pd.read_fwf 的 colspecs 可以這樣用：
      lengths, start_offsets, names = data_rule(table_name)
      colspecs = list(zip(start_offsets[:-1], start_offsets[1:]))
    """
    master_df = _load_master_sheet()
    matched = master_df.loc[master_df["Hadoop \nTable Name"] == target_table_name]
    if matched.empty:
        raise ValueError(f"❌ 檔規總表裡找不到 {target_table_name} 對應的分頁")
    target_sheet_name = matched.iat[0, 1]

    spec_df = pd.read_excel(PATH_DATA_RULE, sheet_name=target_sheet_name)

    use_spec_length_cols = USE_SPEC_LENGTH_COLUMNS.get(target_table_name, set())

    lengths: list[int] = []
    start_offsets: list[int] = [0]
    names: list[str] = []
    cumulative = 0

    for _, row in spec_df.iterrows():
        col_name = row["H_欄位名稱"]
        observed_len = row["S_資料長度"]

        if pd.isnull(observed_len) or pd.isnull(col_name):
            continue

        override_len = FIXED_LENGTH_OVERRIDE.get((target_table_name, col_name))
        if override_len is not None:
            length = override_len
        elif col_name in use_spec_length_cols:
            length = _extract_spec_length(row["H_資料類型長度"])
        else:
            length = int(observed_len)

        cumulative += length
        lengths.append(length)
        start_offsets.append(cumulative)
        names.append(col_name)

    return [lengths, start_offsets, names]
